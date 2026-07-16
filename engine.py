"""
engine.py
Core logic for IDX Securities Financing Portfolio Tracker with Dynamic Excel Formulas.
"""

from __future__ import annotations
import io
import re
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

RATE = 0.095          # 9.5% flat
DAY_BASIS = 360       # 360 basis hari

TEMPLATE_COLUMNS = [
    "CLIENT_ID", "NAME", "B_S", "TRX_DATE", "DUE_DATE", "MATURITY", "ACTIVITY", "STOCK",
    "HC", "VOL", "PRICE", "COLLATERAL_IDR_HC", "AMOUNT_TRX", "FUNDING", "OUTSTANDING", 
    "TRANCHE", "INTEREST", "RATIO", "INV_NO",
]

DISPLAY_HEADERS = {
    "CLIENT_ID": "CLIENT ID",
    "NAME": "NAME",
    "B_S": "B/S",
    "TRX_DATE": "TRX DATE",
    "DUE_DATE": "DUE DATE",
    "MATURITY": "MATURITY",
    "ACTIVITY": "ACTIVITY",
    "STOCK": "STOCK",
    "HC": "HC",
    "VOL": "COLLATERAL (VOL)",
    "PRICE": "PRICE",
    "COLLATERAL_IDR_HC": "COLLATERAL (IDR-HC)",
    "AMOUNT_TRX": "AMOUNT TRX",
    "TRANCHE": "LN (TRANCHE)",
    "FUNDING": "FUNDING",
    "OUTSTANDING": "OUTSTANDING",
    "INTEREST": "INTEREST",
    "RATIO": "RATIO",
    "INV_NO": "INV NO",
}

def _detect_ext(file) -> str:
    name = getattr(file, "name", None)
    if name is None:
        name = str(file)
    return name.lower().rsplit(".", 1)[-1] if "." in name else ""

def _read_table(file, prefer_pipe: bool = False) -> pd.DataFrame:
    ext = _detect_ext(file)
    if hasattr(file, "seek"):
        file.seek(0)
    if ext == "xls":
        return pd.read_excel(file, dtype=str, engine="xlrd")
    if ext == "xlsx":
        return pd.read_excel(file, dtype=str, engine="openpyxl")
    if ext == "txt" or prefer_pipe:
        return pd.read_csv(file, sep="|", dtype=str)
    if ext == "csv":
        return pd.read_csv(file, dtype=str)
    
    head = file.read(8) if hasattr(file, "read") else b""
    if hasattr(file, "seek"):
        file.seek(0)
    if head[:2] == b"PK":
        return pd.read_excel(file, dtype=str, engine="openpyxl")
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return pd.read_excel(file, dtype=str, engine="xlrd")
    if prefer_pipe:
        return pd.read_csv(file, sep="|", dtype=str)
    return pd.read_csv(file, dtype=str)

def parse_risk_parameter(file) -> dict:
    df = _read_table(file, prefer_pipe=True)
    df.columns = [c.strip() for c in df.columns]
    df["StockCode"] = df["StockCode"].astype(str).str.strip()
    df["Haircut"] = pd.to_numeric(df["Haircut"], errors="coerce").fillna(0)
    return dict(zip(df["StockCode"], df["Haircut"]))

def parse_closing_price(file) -> dict:
    df = _read_table(file)
    df.columns = [str(c).strip() for c in df.columns]
    if "no_share" not in df.columns or "kurs_now" not in df.columns:
        raise ValueError("File closing price harus punya kolom 'no_share' dan 'kurs_now'.")
    df["no_share"] = df["no_share"].astype(str).str.strip()
    df["kurs_now"] = pd.to_numeric(df["kurs_now"], errors="coerce")
    return dict(zip(df["no_share"], df["kurs_now"]))

def parse_list_invoice(file, hc_map: dict, price_map: dict) -> pd.DataFrame:
    df = _read_table(file)
    df.columns = [c.strip() for c in df.columns]
    required = {"dt_inv", "no_cust", "name", "bors", "no_share", "tot_vol", "rate", "amt_done", "dt_due", "no_inv"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Kolom wajib hilang di list_invoice: {sorted(missing)}")

    out = pd.DataFrame()
    out["CLIENT_ID"] = df["no_cust"].astype(str).str.strip()
    out["NAME"] = df["name"].astype(str).str.strip()
    out["B_S"] = df["bors"].astype(str).str.strip().str.upper()
    out["TRX_DATE"] = pd.to_datetime(df["dt_inv"], dayfirst=True, errors="coerce")
    out["DUE_DATE"] = pd.to_datetime(df["dt_due"], dayfirst=True, errors="coerce")
    out["MATURITY"] = ""
    out["ACTIVITY"] = "TRX"
    out["STOCK"] = df["no_share"].astype(str).str.strip()

    sign = out["B_S"].map({"B": 1, "S": -1}).fillna(1)
    out["HC"] = out["STOCK"].map(hc_map).fillna(0.0)
    out["VOL"] = pd.to_numeric(df["tot_vol"], errors="coerce").fillna(0.0) * sign
    
    price_from_file = out["STOCK"].map(price_map)
    price_fallback = pd.to_numeric(df["rate"], errors="coerce")
    out["PRICE"] = price_from_file.fillna(price_fallback)
    
    out["COLLATERAL_IDR_HC"] = 0.0
    out["AMOUNT_TRX"] = pd.to_numeric(df["amt_done"], errors="coerce").fillna(0.0) * sign
    out["TRANCHE"] = ""
    out["FUNDING"] = 0.0
    out["OUTSTANDING"] = 0.0
    out["INTEREST"] = 0.0
    out["RATIO"] = ""
    out["INV_NO"] = df["no_inv"].astype(str).str.strip()

    return out[TEMPLATE_COLUMNS]

def parse_previous_template(file) -> dict[str, pd.DataFrame]:
    if file is None:
        return {}
    wb = load_workbook(file, data_only=True)
    result: dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row_idx = None
        for i, r in enumerate(rows):
            cells = [str(c).strip() if c is not None else "" for c in r]
            if "CLIENT ID" in cells:
                header_row_idx = i
                break
        if header_row_idx is None:
            continue
        header = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
        col_idx = {h: i for i, h in enumerate(header)}
        
        data_rows = []
        for r in rows[header_row_idx + 1:]:
            ci = col_idx.get("CLIENT ID", 0)
            client_id = r[ci] if len(r) > ci else None
            if client_id in (None, ""):
                break
            data_rows.append(r)
        if not data_rows:
            continue
        
        rec = []
        for r in data_rows:
            def g(col):
                i = col_idx.get(col)
                return r[i] if i is not None and i < len(r) else None
            rec.append({
                "CLIENT_ID": g("CLIENT ID"), "NAME": g("NAME"), "B_S": g("B/S"),
                "TRX_DATE": g("TRX DATE"), "DUE_DATE": g("DUE DATE"), "MATURITY": g("MATURITY"),
                "ACTIVITY": g("ACTIVITY"), "STOCK": g("STOCK"), "HC": g("HC"),
                "VOL": g("COLLATERAL (VOL)"), "PRICE": g("PRICE"), "COLLATERAL_IDR_HC": g("COLLATERAL (IDR-HC)"),
                "AMOUNT_TRX": g("AMOUNT TRX"), "TRANCHE": g("LN (TRANCHE)"), "FUNDING": g("FUNDING"),
                "OUTSTANDING": g("OUTSTANDING"), "INTEREST": g("INTEREST"), "RATIO": g("RATIO"), "INV_NO": g("INV NO"),
            })
        cdf = pd.DataFrame(rec)
        cdf["TRX_DATE"] = pd.to_datetime(cdf["TRX_DATE"], errors="coerce")
        cdf["DUE_DATE"] = pd.to_datetime(cdf["DUE_DATE"], errors="coerce")
        for c in ["HC", "VOL", "PRICE", "COLLATERAL_IDR_HC", "AMOUNT_TRX", "FUNDING", "OUTSTANDING", "INTEREST"]:
            cdf[c] = pd.to_numeric(cdf[c], errors="coerce").fillna(0.0)
        cdf["TRANCHE"] = cdf["TRANCHE"].fillna("").astype(str)
        cdf["RATIO"] = cdf["RATIO"].fillna("").astype(str)
        cdf["INV_NO"] = cdf["INV_NO"].fillna("").astype(str)
        
        cid_val = str(cdf["CLIENT_ID"].iloc[0]) if len(cdf) and cdf["CLIENT_ID"].iloc[0] else sheet_name
        result[cid_val] = cdf[TEMPLATE_COLUMNS]
    return result

def net_transactions(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out_rows = []
    for (cid, stock, trx_date), g in df.groupby(["CLIENT_ID", "STOCK", "TRX_DATE"], dropna=False):
        net_vol = g["VOL"].sum()
        if round(net_vol) == 0:
            continue
        net_amt = g["AMOUNT_TRX"].sum()
        b_s = "B" if net_vol > 0 else "S"
        due_dates = g["DUE_DATE"].dropna()
        due_date = due_dates.mode().iloc[0] if not due_dates.empty else pd.NaT
        hc = g["HC"].iloc[0]
        price = g["PRICE"].iloc[0]
        inv_no = ",".join(sorted(set(g["INV_NO"].astype(str))))
        
        out_rows.append({
            "CLIENT_ID": cid, "NAME": g["NAME"].iloc[0], "B_S": b_s,
            "TRX_DATE": trx_date, "DUE_DATE": due_date, "MATURITY": "",
            "ACTIVITY": "TRX", "STOCK": stock, "HC": hc, "VOL": net_vol, "PRICE": price,
            "COLLATERAL_IDR_HC": 0, "AMOUNT_TRX": net_amt, "TRANCHE": "",
            "FUNDING": 0.0, "OUTSTANDING": 0.0, "INTEREST": 0.0, "RATIO": "", "INV_NO": inv_no
        })
    if not out_rows:
        return pd.DataFrame(columns=TEMPLATE_COLUMNS)
    return pd.DataFrame(out_rows).sort_values(["TRX_DATE", "CLIENT_ID", "STOCK"]).reset_index(drop=True)[TEMPLATE_COLUMNS]

def merge_client_history(old_df: pd.DataFrame | None, new_df: pd.DataFrame) -> pd.DataFrame:
    if old_df is None or old_df.empty:
        combined = new_df.copy()
    else:
        existing_invs = set()
        for s in old_df["INV_NO"].astype(str):
            if s and not pd.isna(s):
                existing_invs.update(s.split(","))
        add_df = new_df[~new_df["INV_NO"].apply(lambda x: any(i in existing_invs for i in str(x).split(",")))]
        combined = pd.concat([old_df, add_df], ignore_index=True)
    return combined.sort_values(["TRX_DATE", "DUE_DATE", "INV_NO"]).reset_index(drop=True)

def _next_letter(used: list[str]) -> str:
    n = len(used)
    if n < 26: return chr(ord("A") + n)
    return chr(ord("A") + (n // 26) - 1) + chr(ord("A") + (n % 26))

def assign_tranches(df: pd.DataFrame) -> pd.DataFrame:
    df = df.sort_values(["TRX_DATE", "DUE_DATE", "INV_NO"]).reset_index(drop=True)
    used_letters: list[str] = []
    balance: dict[str, float] = {}
    for t in df["TRANCHE"]:
        if isinstance(t, str) and t.strip() and t.strip() not in used_letters:
            used_letters.append(t.strip())
            balance[t.strip()] = 0.0
            
    for idx, row in df.iterrows():
        tranche = str(row["TRANCHE"]).strip() if row["TRANCHE"] else ""
        if not tranche:
            if str(row["B_S"]).upper() == "B":
                tranche = _next_letter(used_letters)
                used_letters.append(tranche)
                balance[tranche] = 0.0
            else:
                tranche = None
                for L in used_letters:
                    if balance.get(L, 0.0) > 0.01:
                        tranche = L
                        break
                if tranche is None:
                    tranche = used_letters[-1] if used_letters else _next_letter(used_letters)
                    if tranche not in used_letters: used_letters.append(tranche)
            df.at[idx, "TRANCHE"] = tranche
        balance[tranche] = balance.get(tranche, 0.0) + (pd.to_numeric(row["AMOUNT_TRX"], errors='coerce') or 0.0)
    return df

def compute_ratio_flag(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    net_by_stock = df.groupby("STOCK")["VOL"].sum()
    df["RATIO"] = df.apply(lambda r: "SHORT" if str(r["B_S"]).upper() == "S" and net_by_stock.get(r["STOCK"], 0) < 0 else "", axis=1)
    return df

def process_client(df: pd.DataFrame, as_of_date) -> pd.DataFrame:
    df = assign_tranches(df)
    df = compute_ratio_flag(df)
    return df.sort_values(["TRX_DATE", "DUE_DATE", "INV_NO"]).reset_index(drop=True)

def build_recap(df: pd.DataFrame, hc_map: dict, price_map: dict):
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), 0.0, 0.0, 0.0
        
    tranche_summary = df.groupby("TRANCHE").agg(
        TOTAL_FUNDING=("AMOUNT_TRX", "sum"), 
        LAST_DUE=("DUE_DATE", "max")
    ).reset_index().sort_values("TRANCHE")
    
    stock_pos = df.groupby("STOCK")["VOL"].sum().reset_index()
    stock_pos = stock_pos[stock_pos["VOL"].round(0) != 0]
    stock_pos["HC"] = stock_pos["STOCK"].map(hc_map).fillna(0.0)
    stock_pos["PRICE"] = stock_pos["STOCK"].map(price_map).fillna(0.0)
    stock_pos["COLLATERAL_IDR_HC"] = stock_pos["VOL"] * stock_pos["PRICE"] * (100 - stock_pos["HC"]) / 100
    
    portfolio_total = stock_pos["COLLATERAL_IDR_HC"].sum()
    
    return tranche_summary, stock_pos, portfolio_total, 0.0, 0.0

# --- WRITER EXCEL DENGAN RUMUS/FORMULA MATRIKS HIDUP ---
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)

def write_workbook(client_results: dict[str, dict], as_of_date) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for client_id, res in client_results.items():
        df = res["df"]
        name = res["name"]
        sheet_name = re.sub(r"[\\/*?:\[\]]", "_", str(client_id))[:31]
        ws = wb.create_sheet(sheet_name)

        # Header Info
        ws.cell(row=1, column=1, value="CLIENT ID").font = BOLD
        ws.cell(row=1, column=2, value=str(client_id))
        ws.cell(row=1, column=4, value="NAME").font = BOLD
        ws.cell(row=1, column=5, value=str(name))
        ws.cell(row=2, column=4, value="As of Date:").font = BOLD
        as_of_cell = ws.cell(row=2, column=5, value=pd.Timestamp(as_of_date).to_pydatetime())
        as_of_cell.number_format = "yyyy-mm-dd"

        # Headers
        headers = [DISPLAY_HEADERS.get(c, c) for c in TEMPLATE_COLUMNS]
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=j, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")

        start_row = 5
        end_data_row = start_row + len(df) - 1
        summary_row = end_data_row + 2  # baris "LAST LOAN" (kolom E berisi tanggal cut-off efektif)

        def set_cell(row_i, col_i, value, number_format=None, bold=False):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.border = BORDER
            if number_format:
                cell.number_format = number_format
            if bold:
                cell.font = BOLD
            return cell

        for idx, row in df.iterrows():
            r = start_row + idx
            
            # Kolom 1-13
            set_cell(r, 1, row["CLIENT_ID"])
            set_cell(r, 2, row["NAME"])
            set_cell(r, 3, row["B_S"])
            set_cell(r, 4, pd.Timestamp(row["TRX_DATE"]).to_pydatetime() if pd.notna(row["TRX_DATE"]) else None, "yyyy-mm-dd")
            set_cell(r, 5, pd.Timestamp(row["DUE_DATE"]).to_pydatetime() if pd.notna(row["DUE_DATE"]) else None, "yyyy-mm-dd")
            set_cell(r, 6, f'=IF(C{r}="B",IF(WEEKDAY(DATE(YEAR(E{r}),MONTH(E{r})+3,DAY(E{r})))>6,DATE(YEAR(E{r}),MONTH(E{r})+3,DAY(E{r}))-1,IF(WEEKDAY(DATE(YEAR(E{r}),MONTH(E{r})+3,DAY(E{r})))<2,DATE(YEAR(E{r}),MONTH(E{r})+3,DAY(E{r}))+1,DATE(YEAR(E{r}),MONTH(E{r})+3,DAY(E{r})))),"")', "yyyy-mm-dd")
            set_cell(r, 7, row["ACTIVITY"])
            set_cell(r, 8, row["STOCK"])
            set_cell(r, 9, row["HC"], "#,##0.00")
            set_cell(r, 10, row["VOL"], "#,##0")
            set_cell(r, 11, row["PRICE"], "#,##0")
            set_cell(r, 12, f"=J{r}*K{r}*(100-I{r})/100", "#,##0")
            set_cell(r, 13, row["AMOUNT_TRX"], "#,##0")

            # 14. FUNDING (Kolom N / 14)
            # Rumus disesuaikan: P sekarang adalah kolom Tranche, N adalah kolom Funding
            f_fund = (
                f'=IF(P{r}="","",IF(OR(C{r}="B",C{r}="SW",C{r}="SD"),M{r},'
                f'IF(ABS(SUM(SUMIFS($M$5:M{r},$A$5:A{r},A{r},$C$5:C{r},{"{"}"S","DR","CR"{"}"},$E$5:E{r},E{r})))<'
                f'SUMIFS($N$5:N{r-1},A$5:A{r-1},A{r},$P$5:P{r-1},P{r})+SUMIFS($O$5:O{r-1},A$5:A{r-1},A{r},$P$5:P{r-1},P{r})+'
                f'ABS(SUM(SUMIFS($N$5:N{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r})))+'
                f'ABS(SUM(SUMIFS($O$5:O{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r}))),'
                f'IF(SUM(SUMIFS($M$5:M{r},$A$5:A{r},A{r},$C$5:C{r},{"{"}"S","DR","CR"{"}"},$E$5:E{r},E{r}))+'
                f'SUMIFS($O$5:O{r-1},A$5:A{r-1},A{r},$P$5:P{r-1},P{r})+'
                f'ABS(SUM(SUMIFS($N$5:N{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r})))+'
                f'ABS(SUM(SUMIFS($O$5:O{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r})))>0,'
                f'SUM(SUMIFS($M$5:M{r},$A$5:A{r},A{r},$C$5:C{r},{"{"}"S","DR","CR"{"}"},$E$5:E{r},E{r}))+'
                f'ABS(SUM(SUMIFS($N$5:N{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r})))+'
                f'ABS(SUM(SUMIFS($O$5:O{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r}))),'
                f'SUM(SUMIFS($M$5:M{r},$A$5:A{r},A{r},$C$5:C{r},{"{"}"S","DR","CR"{"}"},$E$5:E{r},E{r}))+'
                f'SUMIFS($O$5:O{r-1},A$5:A{r-1},A{r},$P$5:P{r-1},P{r})+'
                f'ABS(SUM(SUMIFS($N$5:N{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r})))+'
                f'ABS(SUM(SUMIFS($O$5:O{r-1},$A$5:A{r-1},A{r},$C$5:C{r-1},{"{"}"S","DR","CR"{"}"},$E$5:E{r-1},E{r})))),'
                f'-SUMIFS($N$5:N{r-1},A$5:A{r-1},A{r},$P$5:P{r-1},P{r}))))'
            )
            set_cell(r, 14, f_fund if idx > 0 else f'=IF(P{r}="","",M{r})', "#,##0")

            # 15. OUTSTANDING (Kolom O / 15)
            # outstanding sekarang adalah akumulasi Funding (kolom N) berdasarkan Tranche (kolom P)
            set_cell(r, 15, f'=SUMIFS($N$5:N{r},$P$5:P{r},P{r})', "#,##0")

            # 16. LN / TRANCHE (Kolom P / 16)
            set_cell(r, 16, row["TRANCHE"])

            # 17. INTEREST (Kolom Q / 17)
            # Guard: kalau r adalah baris data TERAKHIR di sheet, range MATCH($P{r+1}:$P{end_data_row})
            # menjadi terbalik (start row > end row) karena end_data_row == r itu sendiri.
            # Excel menormalisasi range terbalik ini jadi $P{r}:$P{r+1}, yang keliru memasukkan
            # baris r sendiri sebagai kandidat MATCH -> ketemu "diri sendiri" -> INDEX ambil sel
            # kosong -> (0 - tanggal) jadi angka negatif raksasa -> interest meledak.
            # Fix: kalau r == end_data_row, langsung pakai term fallback (tidak ada baris tranche
            # berikutnya untuk dicari, IFERROR pun sejatinya akan gagal ke situ juga).
            if r < end_data_row:
                interest_term = (
                    f'IFERROR((INDEX($E{r+1}:$E${summary_row},MATCH(P{r},$P{r+1}:$P${end_data_row},0),1)-E{r})'
                    f'*SUMIFS($N$5:N{r},$A$5:$A{r},A{r},$P$5:$P{r},P{r})*9.5%/360,'
                    f'($E${summary_row}-E{r})*SUMIFS($N$5:N{r},$A$5:$A{r},A{r},$P$5:$P{r},P{r})*9.5%/360)'
                )
            else:
                interest_term = (
                    f'($E${summary_row}-E{r})*SUMIFS($N$5:N{r},$A$5:$A{r},A{r},$P$5:$P{r},P{r})*9.5%/360'
                )
            
            formula_interest = (
                f'=IF($E${summary_row}-E{r}<0,0,IF(OR(SUMIFS($Q$4:Q{r-1},$A$4:A{r-1},A{r},$P$4:P{r-1},P{r})>ABS(N{r}),M{r}=N{r}),'
                f'{interest_term},'
                f'-SUMIFS($Q$4:Q{r-1},A$4:A{r-1},A{r},$P$4:P{r-1},P{r})+{interest_term}))+SUM(R{r})'
            )
            set_cell(r, 17, formula_interest, "#,##0")

            # 18-19. RATIO & INV_NO
            set_cell(r, 18, row["RATIO"])
            set_cell(r, 19, row["INV_NO"])

        # =========================================================================
        # ---- REKAPITULASI BARIS BAWAH: PENEMPATAN SEJAJAR KOLOM ATASNYA ----
        # =========================================================================
        r = summary_row
        
        # 1. Baris LAST LOAN (Diposisikan sesuai struktur kolom tabel)
        formula_last_due = f'=IF($E$2<=LOOKUP(2,1/(NOT(ISBLANK(E$5:E{end_data_row}))),E$5:E{end_data_row}),LOOKUP(2,1/(NOT(ISBLANK(E$5:E{end_data_row}))),E$5:E{end_data_row}),$E$2)'
        c_ldue = ws.cell(row=r, column=5, value=formula_last_due)
        c_ldue.number_format = "yyyy-mm-dd"
        c_ldue.font = BOLD
        c_ldue.border = BORDER

        ws.cell(row=r, column=6, value=str(client_id)).font = BOLD
        ws.cell(row=r, column=6).border = BORDER

        ws.cell(row=r, column=7, value=str(name)).font = BOLD
        ws.cell(row=r, column=7).border = BORDER

        ws.cell(row=r, column=10, value=f"=SUM(J5:J{end_data_row})").number_format = "#,##0"
        ws.cell(row=r, column=10).font = BOLD
        ws.cell(row=r, column=10).border = BORDER
        
        ws.cell(row=r, column=12, value=f"=SUM(L5:L{end_data_row})").number_format = "#,##0"
        ws.cell(row=r, column=12).font = BOLD
        ws.cell(row=r, column=12).border = BORDER

        ws.cell(row=r, column=14, value=f"=SUM(N5:N{end_data_row})").number_format = "#,##0"
        ws.cell(row=r, column=14).font = BOLD
        ws.cell(row=r, column=14).border = BORDER
        
        ws.cell(row=r, column=15, value=f"=SUM(O5:O{end_data_row})").number_format = "#,##0"
        ws.cell(row=r, column=15).font = BOLD
        ws.cell(row=r, column=15).border = BORDER
        
        ws.cell(row=r, column=16, value="LAST LOAN").font = BOLD
        ws.cell(row=r, column=16).border = BORDER

        ws.cell(row=r, column=17, value=f"=SUM(Q5:Q{end_data_row})").number_format = "#,##0"
        ws.cell(row=r, column=17).font = BOLD
        ws.cell(row=r, column=17).border = BORDER
        
        ws.cell(row=r, column=18, value=f"=IF(N{r}=0,0,(N{r}+Q{r})/L{r})").number_format = "0.00%"
        ws.cell(row=r, column=18).font = BOLD
        ws.cell(row=r, column=18).border = BORDER

        # 2. Tabel Breakdown Funding dan Saham (Disusun berdampingan sejajar kolom)
        r_recap = r + 3
        
        # Judul Tabel
        ws.cell(row=r_recap, column=8, value="COLLATERAL SAHAM (posisi saat ini)").font = BOLD
        ws.cell(row=r_recap, column=14, value="PELUNASAN FUNDING").font = BOLD
        
        # Sub-Headers Tabel
        r_recap += 1
        headers_saham = [(8, "STOCK"), (9, "HC"), (10, "VOL"), (11, "PRICE"), (12, "COLLATERAL (IDR-HC)")]
        for col_idx, val in headers_saham:
            c = ws.cell(row=r_recap, column=col_idx, value=val)
            c.font = BOLD
            c.border = BORDER

        headers_funding = [(14, "FUNDING"), (15, "TRANCHE"), (16, "INTEREST"), (17, "DUE DATE")]
        for col_idx, val in headers_funding:
            c = ws.cell(row=r_recap, column=col_idx, value=val)
            c.font = BOLD
            c.border = BORDER

        # Start Isi Data
        r_data_start = r_recap + 1
        
        # - Cetak Baris Saham
        r_saham = r_data_start
        stock_df = df.groupby("STOCK").agg({"HC": "first", "VOL": "sum", "PRICE": "first"}).reset_index()
        stock_df["COLLATERAL_IDR"] = stock_df["VOL"] * stock_df["PRICE"] * (100 - stock_df["HC"]) / 100
        for _, s_row in stock_df.iterrows():
            if round(s_row["VOL"]) != 0:
                ws.cell(row=r_saham, column=8, value=s_row["STOCK"]).border = BORDER
                ws.cell(row=r_saham, column=9, value=s_row["HC"] / 100).number_format = "0%"
                ws.cell(row=r_saham, column=9).border = BORDER
                ws.cell(row=r_saham, column=10, value=s_row["VOL"]).number_format = "#,##0"
                ws.cell(row=r_saham, column=10).border = BORDER
                ws.cell(row=r_saham, column=11, value=s_row["PRICE"]).number_format = "#,##0"
                ws.cell(row=r_saham, column=11).border = BORDER
                ws.cell(row=r_saham, column=12, value=s_row["COLLATERAL_IDR"]).number_format = "#,##0"
                ws.cell(row=r_saham, column=12).border = BORDER
                r_saham += 1

        # - Baris Total PORTOFOLIO (sum Collateral IDR-HC)
        last_saham_row = r_saham - 1 if r_saham > r_data_start else r_data_start
        ws.merge_cells(start_row=r_saham, start_column=8, end_row=r_saham, end_column=11)
        c_port = ws.cell(row=r_saham, column=8, value="PORTOFOLIO")
        c_port.font = BOLD
        c_port.alignment = Alignment(horizontal="center")
        c_port.border = BORDER
        for col in (9, 10, 11):
            ws.cell(row=r_saham, column=col).border = BORDER
        c_port_val = ws.cell(row=r_saham, column=12, value=f"=SUM(L{r_data_start}:L{last_saham_row})")
        c_port_val.number_format = "#,##0"
        c_port_val.font = BOLD
        c_port_val.border = BORDER

        # - Cetak Baris Tranche / Pelunasan Funding (FUNDING | TRANCHE | INTEREST | DUE DATE)
        r_tranche = r_data_start
        tranches = sorted({str(t).strip() for t in df["TRANCHE"] if str(t).strip() and not pd.isna(t)})
        for t in tranches:
            f_fund_sum = f'=SUMIFS($N${start_row}:N{end_data_row},$P${start_row}:P{end_data_row},"{t}")'
            f_int_sum = f'=SUMIFS($Q${start_row}:Q{end_data_row},$P${start_row}:P{end_data_row},"{t}")'
            f_due = f'=_xlfn.MAXIFS($E${start_row}:E{end_data_row},$P${start_row}:P{end_data_row},"{t}")'

            ws.cell(row=r_tranche, column=14, value=f_fund_sum).number_format = "#,##0"
            ws.cell(row=r_tranche, column=14).border = BORDER
            ws.cell(row=r_tranche, column=15, value=t).border = BORDER
            ws.cell(row=r_tranche, column=16, value=f_int_sum).number_format = "#,##0"
            ws.cell(row=r_tranche, column=16).border = BORDER
            c_due = ws.cell(row=r_tranche, column=17, value=f_due)
            c_due.number_format = "d-mmm"
            c_due.border = BORDER
            r_tranche += 1

        # - Baris Total Pelunasan Funding
        last_tranche_row = r_tranche - 1 if r_tranche > r_data_start else r_data_start
        r_tranche += 1
        c_tot_fund = ws.cell(row=r_tranche, column=14, value=f"=SUM(N{r_data_start}:N{last_tranche_row})")
        c_tot_fund.number_format = "#,##0"
        c_tot_fund.font = BOLD
        c_tot_fund.border = BORDER
        c_tot_int = ws.cell(row=r_tranche, column=16, value=f"=SUM(P{r_data_start}:P{last_tranche_row})")
        c_tot_int.number_format = "#,##0"
        c_tot_int.font = BOLD
        c_tot_int.border = BORDER

        # Lebar kolom tetap (auto-width dari panjang formula tidak akurat)
        COLUMN_WIDTHS = {
            1: 12, 2: 26, 3: 6, 4: 12, 5: 12, 6: 12, 7: 16, 8: 10, 9: 8,
            10: 14, 11: 10, 12: 18, 13: 15, 14: 15, 15: 15, 16: 10, 17: 15,
            18: 10, 19: 22,
        }
        for col_idx, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Bekukan baris header supaya tetap terlihat saat scroll
        ws.freeze_panes = "A5"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
